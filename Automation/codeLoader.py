#Loads the Base Code and the Obfuscated code into the appropriate sheets
import os
import openpyxl
from openpyxl.styles import Font, Alignment
import string
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.formatting.rule import Rule
from openpyxl.formula import Tokenizer
from openpyxl.worksheet.datavalidation import DataValidation


def initialFormatting(current_worksheet,sheet_type,o_link_row):

#Creates the header row. Inserts the link to the template, fills in light blue, and freezes it
    
    alphabet = string.ascii_uppercase

    fill = PatternFill(fill_type='solid', fgColor='C9DAF8')
    font = Font(size=15, bold=True)
    
    if(sheet_type == 'o'):
        
       
        for i in range(0, len(alphabet), 1):
            letter = alphabet[i]
            
            if(i%2==0):
                current_worksheet[letter+'1'].value = '=[ObfuscationCategorization.xlsx]Template!'+letter+'22'
            current_worksheet[letter+'1'].fill = fill
            current_worksheet[letter+'1'].font = font

    if(sheet_type == 'base'):
        for i in range(0, len(alphabet), 1):
            letter = alphabet[i]
            
            if(i%2==0):
                current_worksheet[letter+'1'].value = '=[ObfuscationCategorization.xlsx]Template!'+letter+'1'
                
            current_worksheet[letter+'1'].fill = fill
            current_worksheet[letter+'1'].font = font

    current_worksheet.freeze_panes = 'A2'

#Creating the dropdowns
    if(sheet_type == 'base'):

        # Define the list of values and their formatting options
        values = ['High Correct', 'Medium Correct', 'Low Correct','High Maybe', 'Medium Maybe','Low Maybe','Low Incorrect','Medium Incorrect', 'High Incorrect','N/A']
        #Colors for the dropdown values# formats = [PatternFill(fgColor='1dad00'), PatternFill(fgColor='2bff00'), PatternFill(fgColor='98fa84'), PatternFill(fgColor='fa7a02'), PatternFill(fgColor='f78b25'), PatternFill(fgColor='ff9b3d'), PatternFill(fgColor='ff6970'), PatternFill(fgColor='fc323d'), PatternFill(fgColor='db000b'),  PatternFill(fgColor='9ca3ad')]

        data_validation = DataValidation(type="list", formula1=f'"{",".join(values)}"')

        current_worksheet.add_data_validation(data_validation)
        
        #For Question 1, i.e. column I
        for row in current_worksheet.iter_rows(min_row=2, max_row=56, min_col=9, max_col=9):
            for cell in row:
                
                data_validation.add(cell)

        #For Question 2, i.e. column Q
        for row in current_worksheet.iter_rows(min_row=2, max_row=56, min_col=17, max_col=17):
            for cell in row:

                data_validation.add(cell)

        #For Question 3, i.e. column Y
        for row in current_worksheet.iter_rows(min_row=2, max_row=56, min_col=25, max_col=25):
            for cell in row:

                data_validation.add(cell)

    elif(sheet_type == 'o'):
        row_num = 2
        #For Question 1, i.e. column I
        for row in current_worksheet.iter_rows(min_row=2, max_row=56, min_col=9, max_col=9):
            for cell in row:
                
                cell.value = '=B'+str(row_num-1)+'!I'+o_link_row
            row_num+=1
        row_num = 2
        #For Question 2, i.e. column Q
        for row in current_worksheet.iter_rows(min_row=2, max_row=56, min_col=17, max_col=17):

            for cell in row:

                cell.value = '=B'+str(row_num-1)+'!Q'+o_link_row
            row_num+=1

        #For Question 3, i.e. column Y
        for row in current_worksheet.iter_rows(min_row=2, max_row=56, min_col=25, max_col=25):

            for cell in row:

                cell.value = '=B'+str(row_num-1)+'!Y'+o_link_row
            row_num+=1

    






root_dir_workbook = 'C:\\Users\\aswin\\OneDrive - Saint Louis University\\Documents\\REU-2023\\REU-Project\\ObfuscationDatabase'
root_dir = '../ObfuscationDatabase'

# Load the Excel spreadsheet
current_workbook = '\\Q3Test.xlsx'#'\\Jurassic_2.xlsx'#'\\Q2Test.xlsx'#'\\ChatGPT_Q1.xlsx'#'\\PaLM_Q1.xlsx'#'\\JTEST.xlsx'##'\\ObfuscationCategorization.xlsx''\\LM2.xlsx'
workbook = openpyxl.load_workbook(root_dir_workbook+ current_workbook)

name_column = 'A'
ob_code_column = 'C'






for root,dirs, files in sorted(os.walk(root_dir)):
    
# #Load Base Code
#     if(root == '../ObfuscationDatabase\Base_Code'):
#         if 'List_Of_Base_Programs' not in workbook.sheetnames:
#             current_worksheet = workbook.create_sheet(title = 'List_Of_Base_Programs')
#         else:
#             current_worksheet = workbook['List_Of_Base_Programs']
       
#         for file in files:
            
#             with open(root+ '\\'+file, 'r') as temp_file:
#                 file_contents = temp_file.read()
                
#                 temp_string = file[1:].split('.')
#                 cell_num = str(int(temp_string[0])+1)
#                 current_worksheet[name_column+cell_num].value = file.split('.')[0]
#                 current_worksheet['E'+cell_num].value = file_contents
#                 current_worksheet[name_column+cell_num].alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
#                 current_worksheet['E'+cell_num].alignment = Alignment(wrapText=True,horizontal='left', vertical='top')

#Load Obfuscations
    if(root.split('\\')[-1][0] == 'O') :
        folder_name = root.split('\\')[-1]
        
        #If obfuscation sheet has not been created yet
        if folder_name not in workbook.sheetnames:
            current_worksheet_o = workbook.create_sheet(title = folder_name)
            
            initialFormatting(current_worksheet_o,"o",str(int(folder_name[1:])+1))

        else:
            current_worksheet_o = workbook[folder_name]
            initialFormatting(current_worksheet_o,"o",str(int(folder_name[1:])+1))
            
            
        #for each file in the folder
        for file in files:
            
            file_extension = os.path.splitext(file)[1]
            if file_extension != ('.cpp'):
                continue
          
            with open(root+ '\\'+file, 'r', encoding='utf-8') as temp_file:
                file_contents = temp_file.read()
                
    # Add obfuscation to O sheet

                #Getting the base code name from the file name
                base_code_name = file.split('.')[0].split('_')[-1]
                
                #setting the row number to be 1 more than the base code name
                row_number = str(int(base_code_name[1:])+1)
                
                #Cell with Obfuscation name
                name_cell = current_worksheet_o[name_column+row_number]

                #Cell with obfuscated code
                o_code_cell = current_worksheet_o[ob_code_column+row_number]

                #Adding the Data to the cells
                name_cell.value = base_code_name
                o_code_cell.value = file_contents

                #Cell formatting
                name_cell.alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
                o_code_cell.alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
                
                
    # Add obfuscation to base code sheet

                #If the base code sheet has not been created yet
                if base_code_name not in workbook.sheetnames:
                    current_worksheet_b = workbook.create_sheet(title = base_code_name)

                    initialFormatting(current_worksheet_b,"base","")


                else:
                    current_worksheet_b = workbook[base_code_name]
                    initialFormatting(current_worksheet_b,"base","")

                    
                    
                
                #setting the row number to be 1 more than the obfuscation name
                row_number = str(int(folder_name[1:])+1)
                
                #Cell where the name of the base code name is
                name_cell = current_worksheet_b[name_column+row_number]

                #Cell where the base code is
                b_code_cell = current_worksheet_b[ob_code_column+row_number]
                
                #Adding the Data to the cells
                name_cell.value = folder_name
                b_code_cell.value = file_contents

                #Formatting cells
                name_cell.alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
                b_code_cell.alignment = Alignment(wrapText=True,horizontal='left', vertical='top')

# Save the changes
workbook.save(root_dir_workbook+current_workbook)

